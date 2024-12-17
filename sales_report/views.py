from django.db.models import Sum, F, ExpressionWrapper, DecimalField
from django.shortcuts import render,get_object_or_404
from django.utils import timezone
from order_management.models import Order
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
from django.db.models import Sum, Count
from django.db.models.functions import TruncMonth, TruncDay
from dateutil.relativedelta import relativedelta
from .sales_utils import generate_invoice
from openpyxl.utils import get_column_letter
from io import BytesIO
import pytz 
from django.contrib import messages
from decimal import Decimal
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph
# --------------Sales Report---------------#


def sales_report(request):
    today = timezone.now().date()
    report_type = request.GET.get('report_type', 'daily')  # Default to daily if no type is selected
    
    # Set date ranges based on report type
    if report_type == 'daily':
        start_date = today
        end_date = today
    elif report_type == 'weekly':
        start_date = today - timedelta(days=7)
        end_date = today
    elif report_type == 'monthly':
        start_date = today - timedelta(days=30)
        end_date = today
    elif report_type == 'custom':
        # Get custom dates from user input
        start_date = request.GET.get('start_date', today - timedelta(days=30))
        end_date = request.GET.get('end_date', today)
        
        # Check if end_date is after today
        if isinstance(end_date, str):
            end_date = timezone.datetime.strptime(end_date, "%Y-%m-%d").date()
        
        if end_date > today:
            # Add an error message if the end date is after today
            messages.error(request, "End date cannot be in the future.", extra_tags="sales_report")
            # Optionally, set the start and end date to today if invalid
            start_date = today
            end_date = today
    else:
        start_date = today - timedelta(days=30)
        end_date = today

    # Convert start_date to a date object if it's a string
    if isinstance(start_date, str):
        start_date = timezone.datetime.strptime(start_date, "%Y-%m-%d").date()

    # Fetch orders within the date range and annotate with totals
    orders = (
        Order.objects.filter(created_at__date__range=[start_date, end_date])
        .annotate(
            original_total=Sum(F('items__price') * F('items__quantity'), output_field=DecimalField()),
            discounted_price=F('final_price'), 
            annotated_total_discount=F('total_discount')  
        )
    )

    # Calculate overall report values
    total_sales_count = orders.count()
    total_order_amount = orders.aggregate(Sum('original_total'))['original_total__sum'] or 0
    total_discount_amount = orders.aggregate(Sum('annotated_total_discount'))['annotated_total_discount__sum'] or 0

    context = {
        'orders': orders,
        'total_sales_count': total_sales_count,
        'total_order_amount': total_order_amount,
        'total_discount_amount': total_discount_amount,  # Add total discount amount to context
        'start_date': start_date,
        'end_date': end_date,
        'report_type': report_type,
    }

    return render(request, 'sales_report/sales_report.html', context)


def generate_pdf_report(request):
    # Query the necessary orders for the report
    orders = Order.objects.all()  # Adjust this query as needed

    # Create an HTTP response with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'

    # Create the PDF document
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    # Add a title
    styles = getSampleStyleSheet()
    title = Paragraph("<b>Sales Report</b>", styles['Title'])
    subtitle = Paragraph("Generated on: " + timezone.now().strftime("%Y-%m-%d %H:%M:%S"), styles['Normal'])
    elements.append(title)
    elements.append(subtitle)
    elements.append(Paragraph("<br/>", styles['Normal']))  # Add spacing

    # Define the table data
    data = [["Order ID", "Status", "Total Amount", "Date"]]
    for order_index, order in enumerate(orders):
        total_amount = order.total_amount()
        discount = order.total_discount
        row_data = [
            order.id, 
            order.status, 
            f"${total_amount:.2f}", 
            order.created_at.strftime("%Y-%m-%d")
        ]
        data.append(row_data)

    # Create the table
    table = Table(data, colWidths=[80, 100, 100, 100, 120])

    # Define the table style
    table.setStyle(TableStyle([
        # Header style
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),  # Dark Blue Header
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # White Text
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Bold Header
        ('FONTSIZE', (0, 0), (-1, 0), 12),  # Larger Font Size
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Center align header text

        # Row styles
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f2f2f2')),  # Default Light Grey
        ('BACKGROUND', (0, 2), (-1, -1), colors.white),  # Alternate White Rows
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),  # Black Text

        # Grid lines
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#336699')),  # Medium Blue Grid

        # Alignment and padding
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Center align all text
        ('FONTSIZE', (0, 1), (-1, -1), 10),  # Regular Font Size
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),  # Padding for header row
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),  # Padding for data rows
    ]))
    elements.append(table)

    # Build the PDF
    doc.build(elements)
    return response


def generate_excel_report(request):
    # Create a workbook and a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Order Report"

    # Add headers
    ws.append(["Order ID", "User Name", "Order Date", "Total Amount", "Status"])

    # Get orders (you can filter by a date range or status if needed)
    orders = Order.objects.all()

    # Write order data to Excel
    for order in orders:
        # Convert datetime to naive (remove timezone)
        order_created_at = order.created_at
        if order_created_at.tzinfo is not None:
            order_created_at = order_created_at.replace(tzinfo=None)  # Make the datetime naive

        order_status = str(order.status)  # Assuming order.status is a model or choice field

        # Append data for each order to the Excel sheet
        ws.append([
            order.id,
            order.user.get_full_name(),  # Assuming you have a `user` related field
            order_created_at,  # Naive datetime
            order.final_price,  # Adjust based on your Order model's field names
            order_status,
        ])

    # Adjust column widths
    for col in range(1, len(ws[1]) + 1):
        column = get_column_letter(col)
        max_length = 0
        for row in ws.iter_rows():
            try:
                if len(str(row[col - 1].value)) > max_length:
                    max_length = len(row[col - 1].value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook to a BytesIO object
    file = BytesIO()
    wb.save(file)
    file.seek(0)

    # Return the response as an Excel file download
    response = HttpResponse(file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="orders_report.xlsx"'

    return response


def download_invoice(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    
    # Generate the invoice
    buffer = generate_invoice(order)

    # Return the PDF as an HTTP response
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="invoice_{order.id}.pdf"'
    return response


